from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse

from sqlalchemy import inspect, select, text

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

try:
    import cloudinary
    import cloudinary.uploader
except Exception:  # pragma: no cover
    cloudinary = None

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise RuntimeError("openpyxl is required. Install dependencies from requirements.txt.") from exc

from app.core.config import settings
from app.db.init_db import initialize_database
from app.db.session import SessionLocal, engine
from app.models.news import NewsPost
from app.models.organization import Contact, Video
from app.models.products import Product, ProductCategory, ProductImage
from app.models.taxonomy import Language, SiteSetting
from app.utils.slug import slugify


CONTENT_SHEET_INDEX = 0
ASSET_SHEET_INDEX = 1

SITE_SETTING_GROUPS: dict[str, tuple[str, str]] = {
    "site_name": ("general", "Company display name"),
    "company_name": ("general", "Company legal name"),
    "site_tagline": ("general", "Company slogan"),
    "company_slogan": ("general", "Company slogan (legacy key)"),
    "company_logo_url": ("general", "Company logo URL"),
    "company_introduction": ("company", "Company introduction"),
    "company_history": ("company", "Company development history"),
    "company_vision": ("company", "Company vision"),
    "company_mission": ("company", "Company mission"),
    "company_core_values": ("company", "Company core values"),
    "company_leadership": ("company", "Company leadership summary"),
    "company_org_chart": ("company", "Company organization chart"),
    "factory_images_json": ("company", "Factory image URLs in JSON array"),
    "factory_address": ("company", "Factory address"),
    "factory_technology": ("company", "Factory technology"),
    "factory_capacity": ("company", "Factory capacity"),
    "factory_certifications": ("company", "Factory certifications"),
    "product_category_name": ("products", "Primary product category name"),
    "product_category_description": ("products", "Primary product category description"),
    "product_category_image_url": ("products", "Primary product category image URL"),
    "product_short_desc": ("products", "Shared product short description"),
    "product_full_desc": ("products", "Shared product full description"),
    "product_size": ("products", "Shared product size"),
    "product_material": ("products", "Shared product material"),
    "product_color": ("products", "Shared product color"),
    "product_use_case": ("products", "Shared product use-case"),
    "product_catalog_pdf_url": ("products", "Shared product catalog URL"),
    "inquiry_template": ("products", "Inquiry form copy and hints"),
    "contact_form_title": ("contact", "Contact form title"),
    "company_address": ("contact", "Company address"),
    "company_email": ("contact", "Company email"),
    "company_phone": ("contact", "Company hotline"),
    "company_map_coordinate": ("contact", "Raw coordinate string"),
    "company_map_url": ("contact", "Google Maps URL from coordinates"),
}


@dataclass
class ProductSeed:
    sku: str
    name: str
    sort_order: int
    gallery_urls: list[str] = field(default_factory=list)


@dataclass
class VideoSeed:
    title: str
    video_url: str
    sort_order: int


@dataclass
class WorkbookPayload:
    settings: dict[str, str]
    category_name: str
    category_description: str
    category_image_url: str | None
    products: list[ProductSeed]
    videos: list[VideoSeed]
    news_title: str
    news_content: str
    news_image_url: str | None
    news_published_at: datetime | None
    contact_address: str
    contact_email: str
    contact_phone: str
    raw_coordinate: str
    map_url: str | None
    latitude: str
    longitude: str


@dataclass
class ImportReport:
    language_code: str
    asset_mode: str
    dry_run: bool
    site_settings_created: int = 0
    site_settings_updated: int = 0
    categories_created: int = 0
    categories_updated: int = 0
    products_created: int = 0
    products_updated: int = 0
    product_images_synced: int = 0
    videos_created: int = 0
    videos_updated: int = 0
    contacts_created: int = 0
    contacts_updated: int = 0
    news_created: int = 0
    news_updated: int = 0
    assets_uploaded: int = 0
    assets_reused: int = 0
    assets_fallback: int = 0
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "language_code": self.language_code,
            "asset_mode": self.asset_mode,
            "dry_run": self.dry_run,
            "site_settings_created": self.site_settings_created,
            "site_settings_updated": self.site_settings_updated,
            "categories_created": self.categories_created,
            "categories_updated": self.categories_updated,
            "products_created": self.products_created,
            "products_updated": self.products_updated,
            "product_images_synced": self.product_images_synced,
            "videos_created": self.videos_created,
            "videos_updated": self.videos_updated,
            "contacts_created": self.contacts_created,
            "contacts_updated": self.contacts_updated,
            "news_created": self.news_created,
            "news_updated": self.news_updated,
            "assets_uploaded": self.assets_uploaded,
            "assets_reused": self.assets_reused,
            "assets_fallback": self.assets_fallback,
            "warnings": self.warnings,
        }


def normalize_key(value: str | None) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    text = text.replace("Đ", "D").replace("đ", "d")
    folded = unicodedata.normalize("NFKD", text)
    folded = "".join(ch for ch in folded if not unicodedata.combining(ch))
    folded = folded.upper()
    folded = re.sub(r"\s+", " ", folded).strip()
    return folded


def parse_content_sheet(content_sheet) -> dict[str, dict[str, Any]]:
    payload: dict[str, dict[str, Any]] = {}
    for row in range(1, content_sheet.max_row + 1):
        key_raw = content_sheet.cell(row=row, column=1).value
        if key_raw is None:
            continue
        key = normalize_key(str(key_raw))
        value_cell = content_sheet.cell(row=row, column=2)
        value = value_cell.value
        link = value_cell.hyperlink.target if value_cell.hyperlink else None
        payload[key] = {"value": value, "link": link}
    return payload


def parse_product_name_map(text: str) -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []
    for raw_line in str(text or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        match = re.match(r"^([A-Za-z0-9.]+)\s*(?:->|=>|[-–—]+>|→)\s*(.+)$", line)
        if not match:
            match = re.match(r"^([A-Za-z0-9.]+)\s+(.+)$", line)
        if not match:
            continue
        sku = match.group(1).strip().upper()
        name = match.group(2).strip(" -")
        if sku and name:
            rows.append((sku, name))
    return rows


def extract_drive_file_id(url: str | None) -> str | None:
    parsed = urlparse(str(url or "").strip())
    if not parsed.netloc:
        return None

    path_match = re.search(r"/d/([a-zA-Z0-9_-]+)", parsed.path)
    if path_match:
        return path_match.group(1)

    if parsed.path.endswith("/uc"):
        query_id = parse_qs(parsed.query).get("id", [])
        if query_id:
            return query_id[0]

    query_id = parse_qs(parsed.query).get("id", [])
    if query_id:
        return query_id[0]
    return None


def to_drive_direct_url(url: str | None) -> str:
    raw = str(url or "").strip()
    if not raw:
        return ""
    file_id = extract_drive_file_id(raw)
    if not file_id:
        return raw
    return f"https://drive.google.com/uc?export=download&id={file_id}"


def parse_asset_sheet(asset_sheet) -> tuple[dict[str, list[str]], list[VideoSeed], list[str]]:
    product_assets: dict[str, list[str]] = {}
    videos: list[VideoSeed] = []
    factory_images: list[str] = []

    for row in range(1, asset_sheet.max_row + 1):
        first_value = str(asset_sheet.cell(row=row, column=1).value or "").strip()
        if not first_value:
            continue

        key = normalize_key(first_value)
        links: list[tuple[str, str]] = []
        for col in range(2, asset_sheet.max_column + 1):
            cell = asset_sheet.cell(row=row, column=col)
            link = cell.hyperlink.target if cell.hyperlink else None
            if not link:
                continue
            title = str(cell.value or "").strip() or f"asset-{row}-{col}"
            links.append((title, link))

        if not links:
            continue

        if key == "VIDEO":
            for index, (title, link) in enumerate(links, start=1):
                videos.append(VideoSeed(title=title, video_url=link, sort_order=index * 10))
            continue

        if key in {"HINH ANH NHA MAY", "HINH ANH NHA MAY "}:
            factory_images.extend([link for _, link in links])
            continue

        product_assets[first_value.upper()] = [link for _, link in links]

    return product_assets, videos, factory_images


def parse_dms_coordinate(raw: str) -> tuple[str, str, str | None]:
    text = str(raw or "").strip()
    if not text:
        return "0", "0", None

    decimal_match = re.findall(r"-?\d+(?:\.\d+)?", text)
    if len(decimal_match) == 2:
        lat = str(float(decimal_match[0]))
        lng = str(float(decimal_match[1]))
        return lat, lng, f"https://www.google.com/maps?q={lat},{lng}"

    pattern = re.compile(
        r"(\d+)[^\d]+(\d+)[^\d]+(\d+(?:\.\d+)?)[^\dNSWE]*([NS])\s+"
        r"(\d+)[^\d]+(\d+)[^\d]+(\d+(?:\.\d+)?)[^\dNSWE]*([EW])",
        re.IGNORECASE,
    )
    match = pattern.search(text)
    if not match:
        return "0", "0", None

    lat_deg, lat_min, lat_sec, lat_dir, lon_deg, lon_min, lon_sec, lon_dir = match.groups()
    lat = float(lat_deg) + float(lat_min) / 60 + float(lat_sec) / 3600
    lon = float(lon_deg) + float(lon_min) / 60 + float(lon_sec) / 3600
    if lat_dir.upper() == "S":
        lat = -lat
    if lon_dir.upper() == "W":
        lon = -lon

    lat_str = str(round(lat, 6))
    lon_str = str(round(lon, 6))
    map_url = f"https://www.google.com/maps?q={lat_str},{lon_str}"
    return lat_str, lon_str, map_url


def get_content_text(content: dict[str, dict[str, Any]], key: str) -> str:
    node = content.get(key) or {}
    value = node.get("value")
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def get_content_link(content: dict[str, dict[str, Any]], key: str) -> str:
    node = content.get(key) or {}
    return str(node.get("link") or "").strip()


def parse_workbook(excel_path: Path) -> WorkbookPayload:
    workbook = load_workbook(excel_path, data_only=True)
    content_sheet = workbook.worksheets[CONTENT_SHEET_INDEX]
    asset_sheet = workbook.worksheets[ASSET_SHEET_INDEX]
    content = parse_content_sheet(content_sheet)

    product_assets, video_rows, factory_images = parse_asset_sheet(asset_sheet)

    product_name_map = parse_product_name_map(get_content_text(content, "MA SAN PHAM"))
    products: list[ProductSeed] = []
    for index, (sku, name) in enumerate(product_name_map, start=1):
        products.append(
            ProductSeed(
                sku=sku,
                name=name,
                sort_order=index * 10,
                gallery_urls=product_assets.get(sku.upper(), []),
            )
        )

    raw_coordinate = get_content_text(content, "GOOGLE MAP")
    latitude, longitude, map_url = parse_dms_coordinate(raw_coordinate)

    news_raw_datetime = (content.get("NGAY DANG") or {}).get("value")
    if isinstance(news_raw_datetime, datetime):
        news_published_at = (
            news_raw_datetime
            if news_raw_datetime.tzinfo is not None
            else news_raw_datetime.replace(tzinfo=timezone.utc)
        )
    else:
        news_published_at = None

    settings_payload: dict[str, str] = {
        "site_name": get_content_text(content, "TEN CONG TY"),
        "company_name": get_content_text(content, "TEN CONG TY"),
        "site_tagline": get_content_text(content, "SLOGAN"),
        "company_slogan": get_content_text(content, "SLOGAN"),
        "company_logo_url": get_content_link(content, "LOGO"),
        "company_introduction": get_content_text(content, "GIOI THIEU"),
        "company_history": get_content_text(content, "LICH SU PHAT TRIEN"),
        "company_vision": get_content_text(content, "TAM NHIN"),
        "company_mission": get_content_text(content, "SU MENH"),
        "company_core_values": get_content_text(content, "GIA TRI COT LOI"),
        "company_leadership": get_content_text(content, "BAN LANH DAO"),
        "company_org_chart": get_content_text(content, "SO DO TO CHUC"),
        "factory_images_json": json.dumps(factory_images, ensure_ascii=False),
        "factory_address": get_content_text(content, "DIA CHI NHA MAY"),
        "factory_technology": get_content_text(content, "CONG NGHE SAN XUAT"),
        "factory_capacity": get_content_text(content, "CONG SUAT"),
        "factory_certifications": get_content_text(content, "CHUNG NHAN"),
        "product_category_name": get_content_text(content, "TEN DANH MUC"),
        "product_category_description": get_content_text(content, "MO TA DANH MUC"),
        "product_category_image_url": get_content_link(content, "ANH DANH MUC"),
        "product_short_desc": get_content_text(content, "MO TA NGAN"),
        "product_full_desc": get_content_text(content, "MO TA CHI TIET"),
        "product_size": get_content_text(content, "KICH THUOC"),
        "product_material": get_content_text(content, "CHAT LIEU"),
        "product_color": get_content_text(content, "MAU SAC"),
        "product_use_case": get_content_text(content, "UNG DUNG"),
        "product_catalog_pdf_url": get_content_link(content, "CATALOGUE PDF"),
        "inquiry_template": get_content_text(content, "INQUIRY BUTTON"),
        "contact_form_title": get_content_text(content, "FROM LIEN HE"),
        "company_address": get_content_text(content, "DIA CHI"),
        "company_email": get_content_text(content, "EMAIL"),
        "company_phone": get_content_text(content, "SDT"),
        "company_map_coordinate": raw_coordinate,
        "company_map_url": map_url or "",
    }

    return WorkbookPayload(
        settings=settings_payload,
        category_name=get_content_text(content, "TEN DANH MUC"),
        category_description=get_content_text(content, "MO TA DANH MUC"),
        category_image_url=get_content_link(content, "ANH DANH MUC") or None,
        products=products,
        videos=video_rows,
        news_title=get_content_text(content, "TIEU DE"),
        news_content=get_content_text(content, "NOI DUNG"),
        news_image_url=get_content_link(content, "ANH") or None,
        news_published_at=news_published_at,
        contact_address=get_content_text(content, "DIA CHI"),
        contact_email=get_content_text(content, "EMAIL"),
        contact_phone=get_content_text(content, "SDT"),
        raw_coordinate=raw_coordinate,
        map_url=map_url,
        latitude=latitude,
        longitude=longitude,
    )


def has_cloudinary_config() -> bool:
    return bool(
        settings.cloudinary_url.strip()
        or (
            settings.cloudinary_cloud_name.strip()
            and settings.cloudinary_api_key.strip()
            and settings.cloudinary_api_secret.strip()
        )
    )


class AssetResolver:
    def __init__(self, mode: str, report: ImportReport, dry_run: bool) -> None:
        self.mode = mode
        self.report = report
        self.dry_run = dry_run
        self.cache: dict[str, str] = {}
        self.root_folder = f"{settings.cloudinary_folder.strip('/')}/from-web".strip("/")

        if self.mode != "cloudinary":
            return
        if cloudinary is None:
            raise RuntimeError("cloudinary package is not available in this environment.")

        if settings.cloudinary_url.strip():
            cloudinary.config(cloudinary_url=settings.cloudinary_url, secure=True)
            return

        cloudinary.config(
            cloud_name=settings.cloudinary_cloud_name,
            api_key=settings.cloudinary_api_key,
            api_secret=settings.cloudinary_api_secret,
            secure=True,
        )

    def _public_id(self, source_url: str, label: str) -> str:
        drive_id = extract_drive_file_id(source_url)
        if drive_id:
            return drive_id
        label_slug = slugify(label or "asset", fallback="asset")
        digest = hashlib.sha1(source_url.encode("utf-8")).hexdigest()[:10]
        return f"{label_slug}-{digest}"

    def resolve(self, source_url: str | None, *, label: str, group: str, resource_type: str = "auto") -> str | None:
        normalized_source = str(source_url or "").strip()
        if not normalized_source:
            return None

        direct_source = to_drive_direct_url(normalized_source)
        if direct_source in self.cache:
            self.report.assets_reused += 1
            return self.cache[direct_source]

        if self.mode != "cloudinary" or self.dry_run:
            self.cache[direct_source] = direct_source
            return direct_source

        asset_folder = f"{self.root_folder}/{group.strip('/')}".strip("/")
        public_id = self._public_id(direct_source, label)
        try:
            upload_result = cloudinary.uploader.upload(
                direct_source,
                resource_type=resource_type,
                asset_folder=asset_folder,
                public_id=public_id,
                overwrite=True,
                use_filename=False,
                unique_filename=False,
                use_asset_folder_as_public_id_prefix=True,
                display_name=label[:255],
            )
            resolved_url = upload_result.get("secure_url") or upload_result.get("url") or direct_source
            self.cache[direct_source] = resolved_url
            self.report.assets_uploaded += 1
            return resolved_url
        except Exception as exc:
            self.report.assets_fallback += 1
            self.report.warnings.append(f"Asset upload fallback for '{label}': {exc}")
            self.cache[direct_source] = direct_source
            return direct_source


def resolve_asset_mode(mode: str) -> str:
    normalized = mode.strip().lower()
    if normalized in {"cloudinary", "source"}:
        return normalized
    if settings.media_storage.strip().lower() == "cloudinary" and has_cloudinary_config():
        return "cloudinary"
    return "source"


def get_or_create_language(session, language_code: str) -> Language:
    normalized_code = language_code.strip().lower()
    language = session.scalar(select(Language).where(Language.code == normalized_code))
    if language:
        return language

    fallback = session.scalar(select(Language).where(Language.is_default.is_(True)).order_by(Language.id.asc()))
    if fallback:
        return fallback

    language = Language(code=normalized_code, name=normalized_code.upper(), is_default=True, status="active")
    session.add(language)
    session.flush()
    return language


def upsert_site_settings(session, language_id: int, values: dict[str, str], report: ImportReport) -> None:
    existing = {item.config_key: item for item in session.scalars(select(SiteSetting)).all()}
    now = datetime.now(timezone.utc)
    for config_key, config_value in values.items():
        group_name, description = SITE_SETTING_GROUPS.get(config_key, ("import", "Imported from Excel"))
        record = existing.get(config_key)
        if record:
            record.config_value = config_value
            record.language_id = language_id
            record.group_name = group_name
            record.description = description
            record.updated_at = now
            session.add(record)
            report.site_settings_updated += 1
            continue
        session.add(
            SiteSetting(
                config_key=config_key,
                config_value=config_value,
                language_id=language_id,
                group_name=group_name,
                description=description,
                updated_at=now,
            )
        )
        report.site_settings_created += 1


def upsert_category(
    session,
    *,
    name: str,
    description: str,
    image_url: str | None,
    report: ImportReport,
) -> ProductCategory:
    category_slug = slugify(name, fallback="product-category")
    category = session.scalar(select(ProductCategory).where(ProductCategory.slug == category_slug))
    if category is None:
        category = session.scalar(select(ProductCategory).where(ProductCategory.name == name))

    if category is None:
        category = ProductCategory(
            name=name,
            slug=category_slug,
            description=description,
            image_url=image_url,
            sort_order=10,
            is_active=True,
        )
        session.add(category)
        session.flush()
        report.categories_created += 1
        return category

    category.name = name
    category.slug = category_slug
    category.description = description
    category.image_url = image_url
    category.is_active = True
    if not category.sort_order:
        category.sort_order = 10
    session.add(category)
    session.flush()
    report.categories_updated += 1
    return category


def upsert_products(
    session,
    *,
    category: ProductCategory,
    payload: WorkbookPayload,
    report: ImportReport,
) -> None:
    products_by_sku = {str(item.sku or "").upper(): item for item in session.scalars(select(Product)).all() if item.sku}
    products_by_slug = {item.slug: item for item in session.scalars(select(Product)).all()}

    size_value = clip_varchar(payload.settings.get("product_size"), 255)
    material_value = clip_varchar(payload.settings.get("product_material"), 255)
    color_value = clip_varchar(payload.settings.get("product_color"), 255)

    for seed in payload.products:
        sku = seed.sku.upper()
        slug = slugify(f"{seed.sku}-{seed.name}", fallback=f"product-{sku.lower()}")
        record = products_by_sku.get(sku) or products_by_slug.get(slug)
        if record is None:
            record = Product(
                category_id=category.id,
                sku=seed.sku,
                name=seed.name,
                slug=slug,
                short_desc=payload.settings.get("product_short_desc"),
                full_desc=payload.settings.get("product_full_desc"),
                size=size_value,
                material=material_value,
                color=color_value,
                use_case=payload.settings.get("product_use_case"),
                catalog_pdf_url=payload.settings.get("product_catalog_pdf_url"),
                image_url=seed.gallery_urls[0] if seed.gallery_urls else None,
                sort_order=seed.sort_order,
                is_active=True,
            )
            session.add(record)
            session.flush()
            report.products_created += 1
        else:
            record.category_id = category.id
            record.sku = seed.sku
            record.name = seed.name
            record.slug = slug
            record.short_desc = payload.settings.get("product_short_desc")
            record.full_desc = payload.settings.get("product_full_desc")
            record.size = size_value
            record.material = material_value
            record.color = color_value
            record.use_case = payload.settings.get("product_use_case")
            record.catalog_pdf_url = payload.settings.get("product_catalog_pdf_url")
            record.image_url = seed.gallery_urls[0] if seed.gallery_urls else None
            record.sort_order = seed.sort_order
            record.is_active = True
            session.add(record)
            session.flush()
            report.products_updated += 1

        record.images.clear()
        for index, image_url in enumerate(seed.gallery_urls, start=1):
            record.images.append(ProductImage(url=image_url, alt=seed.name, sort_order=index))
            report.product_images_synced += 1
        session.add(record)


def clip_varchar(value: str | None, limit: int) -> str | None:
    if value is None:
        return None
    text = str(value)
    if len(text) <= limit:
        return text
    return text[: max(0, limit - 3)].rstrip() + "..."


def upsert_videos(
    session,
    *,
    language_id: int,
    videos: list[VideoSeed],
    report: ImportReport,
) -> None:
    existing_by_title = {
        str(item.title or "").strip().lower(): item
        for item in session.scalars(select(Video)).all()
        if str(item.title or "").strip()
    }
    for seed in videos:
        key = seed.title.strip().lower()
        record = existing_by_title.get(key)
        if record is None:
            record = Video(
                title=seed.title,
                description=f"{seed.title} (imported from From_Web.xlsx)",
                video_url=seed.video_url,
                language_id=language_id,
                sort_order=seed.sort_order,
                status="published",
            )
            session.add(record)
            report.videos_created += 1
            continue
        record.description = f"{seed.title} (imported from From_Web.xlsx)"
        record.video_url = seed.video_url
        record.language_id = language_id
        record.sort_order = seed.sort_order
        record.status = "published"
        session.add(record)
        report.videos_updated += 1


def upsert_contact(
    session,
    *,
    language_id: int,
    payload: WorkbookPayload,
    company_name: str,
    report: ImportReport,
) -> None:
    contact = session.scalar(
        select(Contact).where(
            Contact.language_id == language_id,
            Contact.is_primary.is_(True),
        )
    )
    if contact is None:
        contact = Contact(
            name=company_name or "Head Office",
            contact_type="head_office",
            address=payload.contact_address,
            phone=payload.contact_phone,
            email=payload.contact_email,
            map_url=payload.map_url,
            latitude=payload.latitude,
            longitude=payload.longitude,
            is_primary=True,
            language_id=language_id,
        )
        session.add(contact)
        report.contacts_created += 1
        return

    contact.name = company_name or contact.name
    contact.contact_type = "head_office"
    contact.address = payload.contact_address
    contact.phone = payload.contact_phone
    contact.email = payload.contact_email
    contact.map_url = payload.map_url
    contact.latitude = payload.latitude
    contact.longitude = payload.longitude
    contact.is_primary = True
    contact.language_id = language_id
    session.add(contact)
    report.contacts_updated += 1


def upsert_news(session, *, payload: WorkbookPayload, report: ImportReport) -> None:
    if not payload.news_title.strip():
        report.warnings.append("News row is empty; skipped news_posts import.")
        return

    slug = slugify(payload.news_title, fallback="news-post")
    news = session.scalar(select(NewsPost).where(NewsPost.slug == slug))
    summary = payload.news_content.strip()[:300] if payload.news_content else None
    if news is None:
        news = NewsPost(
            title=payload.news_title,
            slug=slug,
            summary=summary,
            content=payload.news_content,
            thumbnail_url=payload.news_image_url,
            author=payload.settings.get("company_name"),
            status="published",
            is_featured=True,
            published_at=payload.news_published_at or datetime.now(timezone.utc),
            meta_title=payload.news_title,
            meta_description=summary,
            sort_order=10,
        )
        session.add(news)
        report.news_created += 1
        return

    news.title = payload.news_title
    news.summary = summary
    news.content = payload.news_content
    news.thumbnail_url = payload.news_image_url
    news.author = payload.settings.get("company_name")
    news.status = "published"
    news.is_featured = True
    news.published_at = payload.news_published_at or news.published_at or datetime.now(timezone.utc)
    news.meta_title = payload.news_title
    news.meta_description = summary
    news.sort_order = 10
    session.add(news)
    report.news_updated += 1


def apply_asset_resolution(payload: WorkbookPayload, resolver: AssetResolver) -> WorkbookPayload:
    resolved_settings = dict(payload.settings)

    resolved_settings["company_logo_url"] = resolver.resolve(
        payload.settings.get("company_logo_url"),
        label="company-logo",
        group="branding",
        resource_type="image",
    ) or ""
    resolved_settings["product_category_image_url"] = resolver.resolve(
        payload.settings.get("product_category_image_url"),
        label="product-category-image",
        group="products/category",
        resource_type="image",
    ) or ""
    resolved_settings["product_catalog_pdf_url"] = resolver.resolve(
        payload.settings.get("product_catalog_pdf_url"),
        label="catalogue-pdf",
        group="products/catalog",
        resource_type="raw",
    ) or ""

    factory_images_raw = payload.settings.get("factory_images_json") or "[]"
    try:
        factory_urls = json.loads(factory_images_raw)
    except Exception:
        factory_urls = []
    resolved_factory_urls = []
    for index, url in enumerate(factory_urls, start=1):
        resolved = resolver.resolve(
            url,
            label=f"factory-image-{index}",
            group="factory",
            resource_type="image",
        )
        if resolved:
            resolved_factory_urls.append(resolved)
    resolved_settings["factory_images_json"] = json.dumps(resolved_factory_urls, ensure_ascii=False)

    resolved_products: list[ProductSeed] = []
    for product in payload.products:
        gallery = []
        for index, url in enumerate(product.gallery_urls, start=1):
            resolved = resolver.resolve(
                url,
                label=f"{product.sku}-image-{index}",
                group=f"products/{slugify(product.sku, fallback='sku')}",
                resource_type="image",
            )
            if resolved:
                gallery.append(resolved)
        resolved_products.append(
            ProductSeed(
                sku=product.sku,
                name=product.name,
                sort_order=product.sort_order,
                gallery_urls=gallery,
            )
        )

    resolved_videos: list[VideoSeed] = []
    for index, video in enumerate(payload.videos, start=1):
        resolved_url = resolver.resolve(
            video.video_url,
            label=video.title or f"video-{index}",
            group="videos",
            resource_type="video",
        )
        if not resolved_url:
            continue
        resolved_videos.append(VideoSeed(title=video.title, video_url=resolved_url, sort_order=video.sort_order))

    resolved_news_image = resolver.resolve(
        payload.news_image_url,
        label="news-thumbnail",
        group="news",
        resource_type="image",
    )

    resolved_category_image = resolved_settings.get("product_category_image_url") or None

    return WorkbookPayload(
        settings=resolved_settings,
        category_name=payload.category_name,
        category_description=payload.category_description,
        category_image_url=resolved_category_image,
        products=resolved_products,
        videos=resolved_videos,
        news_title=payload.news_title,
        news_content=payload.news_content,
        news_image_url=resolved_news_image,
        news_published_at=payload.news_published_at,
        contact_address=payload.contact_address,
        contact_email=payload.contact_email,
        contact_phone=payload.contact_phone,
        raw_coordinate=payload.raw_coordinate,
        map_url=payload.map_url,
        latitude=payload.latitude,
        longitude=payload.longitude,
    )


def run(args: argparse.Namespace) -> ImportReport:
    excel_path = Path(args.excel).resolve()
    if not excel_path.exists():
        raise RuntimeError(f"Excel file not found: {excel_path}")

    initialize_database()
    ensure_import_schema()
    workbook_payload = parse_workbook(excel_path)
    asset_mode = resolve_asset_mode(args.asset_mode)
    report = ImportReport(language_code=args.language_code, asset_mode=asset_mode, dry_run=args.dry_run)
    resolver = AssetResolver(mode=asset_mode, report=report, dry_run=args.dry_run)
    resolved_payload = apply_asset_resolution(workbook_payload, resolver)

    with SessionLocal() as session:
        language = get_or_create_language(session, args.language_code)
        upsert_site_settings(session, language_id=language.id, values=resolved_payload.settings, report=report)

        category = upsert_category(
            session,
            name=resolved_payload.category_name,
            description=resolved_payload.category_description,
            image_url=resolved_payload.category_image_url,
            report=report,
        )

        upsert_products(session, category=category, payload=resolved_payload, report=report)
        upsert_videos(session, language_id=language.id, videos=resolved_payload.videos, report=report)
        upsert_contact(
            session,
            language_id=language.id,
            payload=resolved_payload,
            company_name=resolved_payload.settings.get("company_name", ""),
            report=report,
        )
        upsert_news(session, payload=resolved_payload, report=report)

        if args.dry_run:
            session.rollback()
        else:
            session.commit()

    return report


def ensure_import_schema() -> None:
    # Some historical databases still have old varchar(255) columns for long product text fields.
    # Relax them before importing long Vietnamese descriptions from Excel.
    statements = [
        "ALTER TABLE products ALTER COLUMN short_desc TYPE TEXT",
        "ALTER TABLE products ALTER COLUMN full_desc TYPE TEXT",
        "ALTER TABLE products ALTER COLUMN use_case TYPE TEXT",
        "ALTER TABLE products ALTER COLUMN catalog_pdf_url TYPE VARCHAR(2000)",
        "ALTER TABLE products ALTER COLUMN image_url TYPE VARCHAR(1000)",
        "ALTER TABLE products ALTER COLUMN video_url TYPE VARCHAR(2000)",
        "ALTER TABLE product_images ALTER COLUMN url TYPE VARCHAR(2000)",
        "ALTER TABLE news_posts ALTER COLUMN thumbnail_url TYPE VARCHAR(1000)",
    ]
    with engine.begin() as connection:
        inspector = inspect(connection)
        table_names = set(inspector.get_table_names())
        if "products" not in table_names:
            return
        for statement in statements:
            target_table = statement.split()[2]
            if target_table not in table_names:
                continue
            try:
                connection.execute(text(statement))
            except Exception:
                # Ignore unsupported statements on non-Postgres backends.
                pass


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Import From_Web.xlsx into DB and optionally sync Drive assets to Cloudinary."
    )
    parser.add_argument(
        "--excel",
        default=str(PROJECT_ROOT.parent / "China_Web_FE" / "docs" / "From_Web.xlsx"),
        help="Path to From_Web.xlsx",
    )
    parser.add_argument(
        "--language-code",
        default="en",
        help="Language code for imported records (default: en)",
    )
    parser.add_argument(
        "--asset-mode",
        choices=["auto", "cloudinary", "source"],
        default="auto",
        help="Asset target: auto (default), cloudinary, or source (Drive direct links).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Run importer without committing DB changes and without uploading assets.",
    )
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    report = run(args)
    print(json.dumps(report.to_dict(), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
